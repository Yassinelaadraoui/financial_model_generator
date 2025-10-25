"""
fetch_financials_alpha_v2.py

Fetch quarterly and annual financial data for a stock ticker using Alpha Vantage API,
convert to billions, and export to Excel with proper formatting and definitions.

Usage:
    python generator.py
"""

import requests  # Library to make HTTP requests (to call the API)
import pandas as pd  # Library for data manipulation and analysis (using DataFrames)
from openpyxl import load_workbook  # Library to load and modify an existing Excel workbook
from openpyxl.utils import get_column_letter  # Utility to convert a column index (1) to a letter ('A')
from openpyxl.chart import LineChart, Reference  # Import chart components

# Define the base URL for the Alpha Vantage API
ALPHA_VANTAGE_API_URL = "https://www.alphavantage.co/query"


def fetch_alpha_vantage_data(function: str, symbol: str, api_key: str):
    """
    Fetches data from the Alpha Vantage API for a specific function and symbol.
    Includes error checking for common API-level messages.
    """
    # Prepare the query parameters (e.g., ?function=INCOME_STATEMENT&symbol=AAPL...)
    params = {"function": function, "symbol": symbol, "apikey": api_key}
    
    # Send an HTTP GET request to the API URL with the specified parameters
    response = requests.get(ALPHA_VANTAGE_API_URL, params=params)
    
    # Check if the request was successful (e.g., 404, 500)
    response.raise_for_status()
    
    # Parse the JSON response text into a Python dictionary
    json_data = response.json()
    
    # --- NEW: Check for API-level errors in the JSON response ---
    if "Error Message" in json_data:
        raise Exception(f"API Error for {function} {symbol}: {json_data['Error Message']}")
    if "Information" in json_data:
        raise Exception(f"API Info for {function} {symbol}: {json_data['Information']} (This often means you've hit a rate limit)")
    if not json_data: # Handle empty JSON response {}
        raise Exception(f"API Error for {function} {symbol}: Received empty response.")

    # Return the dictionary
    return json_data


def safe_convert_to_float(data_dict: dict, key: str):
    """
    Safely gets a value from a dictionary, converts it to float.
    Returns None if the key doesn't exist or if conversion fails.
    """
    # Get the value associated with the key; returns None if key doesn't exist
    value = data_dict.get(key)
    try:
        # Attempt to convert the retrieved value to a float
        return float(value)
    except (ValueError, TypeError):
        # If conversion fails (e.g., value is "None", None, or a non-numeric string), return None
        return None


def build_quarterly_dataframe(ticker: str, api_key: str) -> pd.DataFrame:
    """
    Fetches, combines, and transforms quarterly income, balance sheet,
    and cash flow data into a single formatted DataFrame.
    """
    # Fetch raw data for all three financial statements
    income_data = fetch_alpha_vantage_data("INCOME_STATEMENT", ticker, api_key)
    balance_sheet_data = fetch_alpha_vantage_data("BALANCE_SHEET", ticker, api_key)
    cash_flow_data = fetch_alpha_vantage_data("CASH_FLOW", ticker, api_key)

    # Extract the list of quarterly reports from the JSON response
    # Default to an empty list ([]) if the 'quarterlyReports' key is missing
    income_reports = income_data.get("quarterlyReports", [])
    balance_sheet_reports = balance_sheet_data.get("quarterlyReports", [])
    cash_flow_reports = cash_flow_data.get("quarterlyReports", [])
    
    # This list will hold a dictionary for each quarter's combined data
    quarterly_data_rows = []

    # Iterate through the reports. We use min() to avoid an IndexError
    # if one statement has fewer reports than the others.
    # This assumes reports are in the same chronological order (newest first).
    num_reports = min(len(income_reports), len(balance_sheet_reports), len(cash_flow_reports))
    
    # If num_reports is 0, print the warning
    if num_reports == 0:
        print("⚠️ Warning: No quarterly data fetched. Output may be incomplete.")
        return pd.DataFrame() # Return an empty DataFrame
        
    for i in range(num_reports):
        # Get the specific report dictionaries for the current period (index i)
        income_dict = income_reports[i]
        balance_sheet_dict = balance_sheet_reports[i]
        cash_flow_dict = cash_flow_reports[i]

        # Build a single dictionary (a 'row') for the current period.
        # Use safe_convert_to_float to handle 'None' or missing values gracefully.
        row_data = {
            "Date": income_dict.get("fiscalDateEnding"),
            # --- Income Statement Metrics ---
            "Revenue": safe_convert_to_float(income_dict, "totalRevenue"),
            "COGS": safe_convert_to_float(income_dict, "costOfRevenue"),
            "Gross Margin": None,  # Will be calculated next
            "R&D": safe_convert_to_float(income_dict, "researchAndDevelopment"),
            "M&S": None,  # Marketing & Sales is often not broken out separately
            "G&A": safe_convert_to_float(income_dict, "sellingGeneralAndAdministrative"),
            "OpEx": safe_convert_to_float(income_dict, "operatingExpenses"),
            "OpInc": safe_convert_to_float(income_dict, "operatingIncome"),
            "Interest Income": safe_convert_to_float(income_dict, "interestIncome"),
            "Pretax Income": safe_convert_to_float(income_dict, "incomeBeforeTax"),
            "Taxes": safe_convert_to_float(income_dict, "incomeTaxExpense"),
            "Net Income": safe_convert_to_float(income_dict, "netIncome"),
            "EPS": safe_convert_to_float(income_dict, "reportedEPS"),
            "Shares": safe_convert_to_float(income_dict, "commonStockSharesOutstanding"),
            # --- Balance Sheet Metrics ---
            "Cash": safe_convert_to_float(balance_sheet_dict, "cashAndCashEquivalentsAtCarryingValue"),
            "AR": safe_convert_to_float(balance_sheet_dict, "currentNetReceivables"),
            "Prepaids": None,  # Not commonly available in this API breakout
            "PP&E": safe_convert_to_float(balance_sheet_dict, "propertyPlantEquipment"),
            "Lease (Asset)": None, # Not commonly available
            "Goodwill": safe_convert_to_float(balance_sheet_dict, "goodwill"),
            "DTA": None, # Deferred Tax Assets - Not commonly available
            "ONCA": None, # Other Non-Current Assets
            "AP": safe_convert_to_float(balance_sheet_dict, "currentAccountsPayable"),
            "Accrued": None, # Accrued Liabilities - Not commonly available
            "DR": safe_convert_to_float(balance_sheet_dict, "deferredRevenue"),
            "Lease (Liability)": None, # Not commonly available
            "Debt": safe_convert_to_float(balance_sheet_dict, "shortLongTermDebtTotal"),
            "ONCL": None, # Other Non-Current Liabilities
            "SE": safe_convert_to_float(balance_sheet_dict, "totalShareholderEquity"),
            "Assets": safe_convert_to_float(balance_sheet_dict, "totalAssets"),
            # Checksum: Liabilities + Shareholders' Equity (should equal Assets)
            "L+SE": (safe_convert_to_float(balance_sheet_dict, "totalLiabilities") or 0)
            + (safe_convert_to_float(balance_sheet_dict, "totalShareholderEquity") or 0),
            # --- Cash Flow Statement Metrics ---
            "CFFO": safe_convert_to_float(cash_flow_dict, "operatingCashflow"),
            # --- TTM Placeholders (will be calculated later) ---
            "TTM CFFO": None,
            "TTM Revenue": None,
            "TTM Net Income": None,
        }

        # Calculate Gross Margin (as a percentage) if data is available
        if row_data["Revenue"] and row_data["COGS"]:
            row_data["Gross Margin"] = (row_data["Revenue"] - row_data["COGS"]) / row_data["Revenue"]

        # Add the completed dictionary for this quarter to our list of rows
        quarterly_data_rows.append(row_data)

    # Create the initial DataFrame from the list of row dictionaries
    # Each dictionary becomes a row, keys become columns
    quarterly_df = pd.DataFrame(quarterly_data_rows)
    
    # --- TTM (Trailing Twelve Months) Calculations ---
    # We already checked if the df is empty by checking num_reports
    # Sort by date ascending (oldest to newest) to correctly calculate a rolling sum
    quarterly_df = quarterly_df.sort_values(by="Date", ascending=True).reset_index(drop=True)
    
    # Calculate TTM metrics as a 4-quarter rolling sum.
    # `min_periods=4` ensures we only get a value when we have 4 full quarters of data.
    quarterly_df['TTM CFFO'] = quarterly_df['CFFO'].rolling(window=4, min_periods=4).sum()
    quarterly_df['TTM Revenue'] = quarterly_df['Revenue'].rolling(window=4, min_periods=4).sum()
    quarterly_df['TTM Net Income'] = quarterly_df['Net Income'].rolling(window=4, min_periods=4).sum()
    
    # Sort back to newest-first (descending) to match the API's original order
    quarterly_df = quarterly_df.sort_values(by="Date", ascending=False).reset_index(drop=True)

    # --- Transpose DataFrame ---
    # Pivot the table: Dates become columns, metrics become rows
    # 1. Set the 'Date' column as the index
    # 2. Transpose (.T) the DataFrame
    quarterly_df = quarterly_df.set_index("Date").T

    # --- Unit Conversion (to Billions) ---
    # Create a list of row indices (metrics) that need to be divided by 1 billion
    # Exclude metrics that are ratios or per-share values
    metrics_to_convert = [idx_label for idx_label in quarterly_df.index if idx_label not in ["Gross Margin", "EPS"]]
    
    # Iterate over the metric names (row indices)
    for metric in metrics_to_convert:
        try:
            # Convert all values in the row to float and divide by 1e9 (1,000,000,000)
            quarterly_df.loc[metric] = quarterly_df.loc[metric].astype(float) / 1e9
        except Exception:
            # Skip if conversion fails (e.g., row contains non-numeric data)
            pass

    # --- Rename Rows ---
    # Create a mapping (dictionary) for rows that were converted to billions
    index_rename_map = {
        idx_label: f"{idx_label} (B$)" 
        for idx_label in quarterly_df.index 
        if idx_label not in ["Gross Margin", "EPS"]
    }
    
    # Apply the renaming to the DataFrame's index (the row labels)
    quarterly_df.rename(index=index_rename_map, inplace=True)

    # Return the final, formatted DataFrame
    return quarterly_df


def build_annual_dataframe(ticker: str, api_key: str) -> pd.DataFrame:
    """
    Fetches, combines, and transforms *annual* income, balance sheet,
    and cash flow data into a single formatted DataFrame.
    """
    # Fetch raw data for all three financial statements
    income_data = fetch_alpha_vantage_data("INCOME_STATEMENT", ticker, api_key)
    balance_sheet_data = fetch_alpha_vantage_data("BALANCE_SHEET", ticker, api_key)
    cash_flow_data = fetch_alpha_vantage_data("CASH_FLOW", ticker, api_key)

    # Extract the list of *annual* reports from the JSON response
    income_reports = income_data.get("annualReports", [])
    balance_sheet_reports = balance_sheet_data.get("annualReports", [])
    cash_flow_reports = cash_flow_data.get("annualReports", [])
    
    # This list will hold a dictionary for each year's combined data
    annual_data_rows = []

    # Iterate through the reports
    num_reports = min(len(income_reports), len(balance_sheet_reports), len(cash_flow_reports))
    
    # If num_reports is 0, print the warning and return an empty DataFrame
    if num_reports == 0:
        print("⚠️ Warning: No annual data fetched. Output may be incomplete.")
        return pd.DataFrame()
        
    for i in range(num_reports):
        # Get the specific report dictionaries for the current period (index i)
        income_dict = income_reports[i]
        balance_sheet_dict = balance_sheet_reports[i]
        cash_flow_dict = cash_flow_reports[i]

        # Build a single dictionary (a 'row') for the current period
        row_data = {
            "Date": income_dict.get("fiscalDateEnding"),
            # --- Income Statement Metrics ---
            "Revenue": safe_convert_to_float(income_dict, "totalRevenue"),
            "COGS": safe_convert_to_float(income_dict, "costOfRevenue"),
            "Gross Margin": None,  # Will be calculated next
            "R&D": safe_convert_to_float(income_dict, "researchAndDevelopment"),
            "M&S": None,
            "G&A": safe_convert_to_float(income_dict, "sellingGeneralAndAdministrative"),
            "OpEx": safe_convert_to_float(income_dict, "operatingExpenses"),
            "OpInc": safe_convert_to_float(income_dict, "operatingIncome"),
            "Interest Income": safe_convert_to_float(income_dict, "interestIncome"),
            "Pretax Income": safe_convert_to_float(income_dict, "incomeBeforeTax"),
            "Taxes": safe_convert_to_float(income_dict, "incomeTaxExpense"),
            "Net Income": safe_convert_to_float(income_dict, "netIncome"),
            "EPS": safe_convert_to_float(income_dict, "reportedEPS"),
            "Shares": safe_convert_to_float(income_dict, "commonStockSharesOutstanding"),
            # --- Balance Sheet Metrics ---
            "Cash": safe_convert_to_float(balance_sheet_dict, "cashAndCashEquivalentsAtCarryingValue"),
            "AR": safe_convert_to_float(balance_sheet_dict, "currentNetReceivables"),
            "Prepaids": None,
            "PP&E": safe_convert_to_float(balance_sheet_dict, "propertyPlantEquipment"),
            "Lease (Asset)": None,
            "Goodwill": safe_convert_to_float(balance_sheet_dict, "goodwill"),
            "DTA": None,
            "ONCA": None,
            "AP": safe_convert_to_float(balance_sheet_dict, "currentAccountsPayable"),
            "Accrued": None,
            "DR": safe_convert_to_float(balance_sheet_dict, "deferredRevenue"),
            "Lease (Liability)": None,
            "Debt": safe_convert_to_float(balance_sheet_dict, "shortLongTermDebtTotal"),
            "ONCL": None,
            "SE": safe_convert_to_float(balance_sheet_dict, "totalShareholderEquity"),
            "Assets": safe_convert_to_float(balance_sheet_dict, "totalAssets"),
            "L+SE": (safe_convert_to_float(balance_sheet_dict, "totalLiabilities") or 0)
            + (safe_convert_to_float(balance_sheet_dict, "totalShareholderEquity") or 0),
            # --- Cash Flow Statement Metrics ---
            "CFFO": safe_convert_to_float(cash_flow_dict, "operatingCashflow"),
            # --- No TTM calculations for annual data ---
        }

        # Calculate Gross Margin (as a percentage) if data is available
        if row_data["Revenue"] and row_data["COGS"]:
            row_data["Gross Margin"] = (row_data["Revenue"] - row_data["COGS"]) / row_data["Revenue"]

        # Add the completed dictionary for this year to our list of rows
        annual_data_rows.append(row_data)

    # Create the initial DataFrame from the list of row dictionaries
    annual_df = pd.DataFrame(annual_data_rows)

    # --- Transpose DataFrame ---
    annual_df = annual_df.set_index("Date").T

    # --- FIX: Sort columns by date (oldest to newest) for correct chart plotting ---
    annual_df = annual_df.sort_index(axis=1, ascending=True)

    # --- Unit Conversion (to Billions) ---
    metrics_to_convert = [idx_label for idx_label in annual_df.index if idx_label not in ["Gross Margin", "EPS"]]
    
    for metric in metrics_to_convert:
        try:
            annual_df.loc[metric] = annual_df.loc[metric].astype(float) / 1e9
        except Exception:
            pass

    # --- Rename Rows ---
    index_rename_map = {
        idx_label: f"{idx_label} (B$)" 
        for idx_label in annual_df.index 
        if idx_label not in ["Gross Margin", "EPS"]
    }
    
    annual_df.rename(index=index_rename_map, inplace=True)

    # Return the final, formatted DataFrame
    return annual_df


def create_definitions_dataframe() -> pd.DataFrame:
    """Creates a DataFrame containing definitions for key financial metrics."""
    # A dictionary mapping metric names (keys) to their definitions (values)
    definitions_map = {
        "Revenue": "Total sales before any costs are deducted.",
        "COGS": "Cost of goods sold – direct costs of producing goods/services.",
        "Gross Margin": "Profit after COGS, as a percentage of revenue.",
        "R&D": "Research & Development expenses.",
        "M&S": "Marketing & Sales expenses.",
        "G&A": "General & Administrative overhead costs.",
        "OpEx": "Operating expenses (R&D + M&S + G&A).",
        "OpInc": "Operating income (Gross profit – OpEx).",
        "Interest Income": "Income from interest-bearing assets.",
        "Pretax Income": "Earnings before taxes.",
        "Taxes": "Income tax expense.",
        "Net Income": "Final profit after taxes.",
        "EPS": "Earnings per share (Net Income / Shares).",
        "Shares": "Weighted average shares outstanding.",
        "Cash": "Cash and equivalents.",
        "AR": "Accounts receivable – customer balances.",
        "Prepaids": "Prepaid expenses (rent, insurance).",
        "PP&E": "Property, Plant & Equipment.",
        "Lease (Asset)": "Right-of-use leased assets.",
        "Goodwill": "Premium paid on acquisitions.",
        "DTA": "Deferred Tax Assets.",
        "ONCA": "Other Non-Current Assets.",
        "AP": "Accounts Payable.",
        "Accrued": "Accrued but unpaid expenses.",
        "DR": "Deferred Revenue.",
        "Lease (Liability)": "Lease obligations.",
        "Debt": "Short + Long-term debt.",
        "ONCL": "Other Non-Current Liabilities.",
        "SE": "Shareholders’ Equity.",
        "Assets": "Total company assets.",
        "L+SE": "Liabilities + Shareholders’ Equity.",
        "CFFO": "Cash Flow From Operations.",
        "TTM CFFO": "Trailing Twelve Month CFFO.",
        "TTM Revenue": "Trailing Twelve Month Revenue.",
        "TTM Net Income": "Trailing Twelve Month Net Income."
    }
    # Convert the dictionary's .items() (a list of (key, value) tuples) into a DataFrame
    return pd.DataFrame(definitions_map.items(), columns=["Metric", "Definition"])


def _format_financials_sheet(worksheet):
    """
    Helper function to apply formatting (freeze panes, number formats, auto-fit)
    to a given financial data worksheet.
    """
    # --- Freeze Panes ---
    # Freeze top row (dates) and first col (metrics)
    worksheet.freeze_panes = 'B2'

    # --- Apply Formats and Find Max Column Lengths ---
    column_max_lengths = {}

    for row_idx, row in enumerate(worksheet.iter_rows(), 1):
        # Get the first cell (in Col A) which contains the metric name
        metric_name_cell = row[0]
        metric_name = metric_name_cell.value

        # Determine the correct number format based on the metric name in Col A
        target_format = None  # Default
        if metric_name: # Check if metric_name is not None
            if "Gross Margin" in metric_name:
                target_format = '0.0%'  # Percentage format
            elif "EPS" in metric_name:
                target_format = '$#,##0.00'  # Currency format
            elif "(B$)" in metric_name:
                target_format = '#,##0.00'  # Standard number format

        # Iterate over every cell in the current row
        for cell in row:
            col_idx = cell.column  # Get the column index (1, 2, 3...)

            # 1. Apply number formatting
            if row_idx > 1 and col_idx > 1 and isinstance(cell.value, (int, float)):
                if target_format:
                    cell.number_format = target_format

            # 2. Track max content length for auto-fitting columns
            current_max_col_length = column_max_lengths.get(col_idx, 0)
            current_cell_length = len(str(cell.value)) if cell.value is not None else 0
            
            if current_cell_length > current_max_col_length:
                column_max_lengths[col_idx] = current_cell_length

    # --- Auto-fit Columns ---
    for col_idx, max_length in column_max_lengths.items():
        column_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[column_letter].width = max_length + 3


def export_to_excel(quarterly_df: pd.DataFrame, annual_df: pd.DataFrame, definitions_df: pd.DataFrame, ticker: str):
    """Writes the DataFrames to an Excel file and applies custom formatting."""
    # Create the output filename based on the ticker symbol (e.g., "AAPL_financials.xlsx")
    output_filename = f"{ticker.upper()}_financials.xlsx"
    
    # Use pd.ExcelWriter as a context manager to write all DataFrames to the same file
    with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
        # Write the quarterly data
        quarterly_df.to_excel(writer, sheet_name="Quarterly Data", index=True)
        
        # Write the new annual data
        annual_df.to_excel(writer, sheet_name="Annual Data", index=True)
        
        # Write the definitions data
        definitions_df.to_excel(writer, sheet_name="Definitions", index=False)

    # --- Post-Write Formatting using openpyxl ---
    
    # Load the Excel file we just saved
    workbook = load_workbook(output_filename)
    
    # --- Format Financial Sheets ---
    # Get the "Quarterly Data" worksheet object and format it
    if "Quarterly Data" in workbook.sheetnames and not quarterly_df.empty:
        data_worksheet = workbook["Quarterly Data"]
        _format_financials_sheet(data_worksheet)

    # Get the "Annual Data" worksheet object and format it
    if "Annual Data" in workbook.sheetnames and not annual_df.empty:
        annual_worksheet = workbook["Annual Data"]
        _format_financials_sheet(annual_worksheet)

    # --- Format Definitions Sheet ---
    if "Definitions" in workbook.sheetnames:
        definitions_worksheet = workbook["Definitions"]
        # Set a fixed width for the "Metric" column (Column A)
        definitions_worksheet.column_dimensions[get_column_letter(1)].width = 20
        # Set a fixed (wider) width for the "Definition" column (Column B)
        definitions_worksheet.column_dimensions[get_column_letter(2)].width = 50

    # --- NEW: Create Charts Sheet ---
    # Only create charts if the annual_df was not empty
    if "Annual Data" in workbook.sheetnames and not annual_df.empty:
        annual_worksheet = workbook["Annual Data"] # Get the sheet again
        
        # Check if there is data to chart
        if annual_worksheet.max_row <= 1 or annual_worksheet.max_column <= 1:
            print("⚠️ Skipping chart generation: No annual data to plot.")
        else:
            # Create the new charts sheet
            chart_worksheet = workbook.create_sheet("Charts")
            print("📈 Generating charts...")

            # Get dimensions of the annual data
            max_row = annual_worksheet.max_row
            max_col = annual_worksheet.max_column

            # Define the x-axis (categories) - the dates in row 1
            # We now read from oldest (col 2) to newest (max_col)
            categories_ref = Reference(annual_worksheet, min_col=2, min_row=1, max_col=max_col, max_row=1)

            # Loop through each data row (each metric)
            for row_index in range(2, max_row + 1):
                # Create a new line chart
                chart = LineChart()
                
                # Define the data series (the values in the current row)
                data_ref = Reference(annual_worksheet, min_col=2, min_row=row_index, max_col=max_col, max_row=row_index)
                
                # Add the data to the chart
                chart.add_data(data_ref, titles_from_data=False)
                
                # Set the x-axis categories
                chart.set_categories(categories_ref)
                
                # Get the metric name for the chart title
                chart.title = annual_worksheet.cell(row=row_index, column=1).value
                
                # Style the chart
                chart.legend = None # Hide the legend
                chart.y_axis.title = "Value"
                chart.x_axis.title = "Fiscal Year End"

                # Calculate where to place the chart on the sheet
                # This stacks them vertically, with 15 rows per chart
                anchor_cell = f"A{(row_index - 2) * 15 + 1}"
                
                # Add the chart to the "Charts" sheet
                chart_worksheet.add_chart(chart, anchor_cell)

    # Save all the formatting and new charts back to the Excel file
    workbook.save(output_filename)
    
    # Print a success message to the console
    print(f"✅ Exported formatted Excel: {output_filename}")


def main():
    """Main execution function to run the script."""
    print("📊 Fetch quarterly and annual financials (Alpha Vantage)")
    
    # Prompt the user for the stock ticker
    ticker_symbol = input("Ticker (e.g. AAPL): ").strip().upper()
    
    # Prompt the user for their API key
    api_key = input("Alpha Vantage API key: ").strip()

    try:
        # Fetch quarterly data
        print("Fetching quarterly data...")
        quarterly_financials_df = build_quarterly_dataframe(ticker_symbol, api_key)
        
        # Fetch annual data
        print("Fetching annual data...")
        annual_financials_df = build_annual_dataframe(ticker_symbol, api_key)
        
        # Only proceed if at least one of the DataFrames is not empty
        if not quarterly_financials_df.empty or not annual_financials_df.empty:
            # Create the definitions DataFrame
            definitions_df = create_definitions_dataframe()
            # Export all DataFrames to the formatted Excel file
            export_to_excel(quarterly_financials_df, annual_financials_df, definitions_df, ticker_symbol)
        else:
            # Inform the user if no data was processed (e.g., bad ticker)
            print("No data processed for either quarterly or annual reports.")
            
    except Exception as error:
        # Catch any exceptions that occurred
        # With the new error handling, this will now print API limit errors
        print(f"❌ An error occurred: {error}")


# Standard Python entry point
if __name__ == "__main__":
    main()