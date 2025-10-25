"""
excel_exporter.py

Contains all logic for exporting DataFrames to a
formatted Excel workbook, including all styling and charting.
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles.fills import PatternFill

def _format_financials_sheet(worksheet):
    """
    Helper function to apply formatting (freeze panes, number formats, auto-fit)
    to a given financial data worksheet.
    """
    worksheet.freeze_panes = 'B2'

    format_percent = '0.0%'
    format_currency = '$#,##0.00'
    format_billion = '#,##0.00'
    format_ratio = '0.00'
    
    red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    column_max_lengths = {}
    for row_idx, row in enumerate(worksheet.iter_rows(), 1):
        metric_name_cell = row[0]
        metric_name = metric_name_cell.value
        target_format = None
        
        if metric_name:
            if "Margin" in metric_name or "ROE" in metric_name or "ROA" in metric_name or "Growth" in metric_name or "Yield" in metric_name or "Payout" in metric_name:
                target_format = format_percent
            elif "EPS" in metric_name:
                target_format = format_currency
            elif "(B$)" in metric_name:
                target_format = format_billion
            elif "Ratio" in metric_name or "Coverage" in metric_name or "Debt-to-Equity" in metric_name or "PEGRatio" in metric_name or "EVToEBITDA" in metric_name:
                target_format = format_ratio
        
        for cell in row:
            col_idx = cell.column
            
            if row_idx > 1 and col_idx > 1 and isinstance(cell.value, (int, float)):
                if target_format:
                    cell.number_format = target_format
            
            current_max_col_length = column_max_lengths.get(col_idx, 0)
            current_cell_length = len(str(cell.value)) if cell.value is not None else 0
            if current_cell_length > current_max_col_length:
                column_max_lengths[col_idx] = current_cell_length
                
        if metric_name and ("Growth" in metric_name or "Free Cash Flow" in metric_name):
            cell_range = f"B{row_idx}:{get_column_letter(worksheet.max_column)}{row_idx}"
            worksheet.conditional_formatting.add(cell_range,
                CellIsRule(operator='lessThan', formula=[0], fill=red_fill))

    for col_idx, max_length in column_max_lengths.items():
        column_letter = get_column_letter(col_idx)
        worksheet.column_dimensions[column_letter].width = max_length + 3


def _add_charts_to_sheet(worksheet, data_sheet, metrics_to_chart):
    """
    Helper function to create charts on the "Charts" sheet
    based on data from the "Annual Data" sheet.
    """
    if data_sheet.max_row <= 1 or data_sheet.max_column <= 1:
        print("⚠️ Skipping chart generation: No annual data to plot.")
        return

    max_col = data_sheet.max_column
    categories_ref = Reference(data_sheet, min_col=2, min_row=1, max_col=max_col, max_row=1)
    chart_row_anchor = 1
    
    metric_row_map = {}
    for row_idx in range(2, data_sheet.max_row + 1):
        metric_row_map[data_sheet.cell(row=row_idx, column=1).value] = row_idx

    for metric_name in metrics_to_chart:
        row_index = metric_row_map.get(metric_name)
        
        if not row_index:
            print(f"⚠️ Could not find metric '{metric_name}' for charting.")
            continue
            
        chart = LineChart()
        data_ref = Reference(data_sheet, min_col=2, min_row=row_index, max_col=max_col, max_row=row_index)
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(categories_ref)
        chart.title = metric_name
        chart.legend = None
        chart.y_axis.title = "Value"
        chart.x_axis.title = "Fiscal Year End"
        
        anchor_cell = f"A{chart_row_anchor}"
        worksheet.add_chart(chart, anchor_cell)
        chart_row_anchor += 15


def export_to_excel(ticker: str, summary_df: pd.DataFrame, quarterly_df: pd.DataFrame, annual_df: pd.DataFrame, price_df: pd.DataFrame, definitions_df: pd.DataFrame):
    """Writes all DataFrames to a formatted Excel file."""
    output_filename = f"{ticker.upper()}_financials.xlsx"
    
    with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="Summary", index=True)
        quarterly_df.to_excel(writer, sheet_name="Quarterly Data", index=True)
        annual_df.to_excel(writer, sheet_name="Annual Data", index=True)
        price_df.to_excel(writer, sheet_name="Stock Price", index=True)
        definitions_df.to_excel(writer, sheet_name="Definitions", index=False)

    workbook = load_workbook(output_filename)
    
    if "Summary" in workbook.sheetnames:
        ws = workbook["Summary"]
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        _format_financials_sheet(ws)
    
    if "Quarterly Data" in workbook.sheetnames and not quarterly_df.empty:
        ws = workbook["Quarterly Data"]
        _format_financials_sheet(ws)

    if "Annual Data" in workbook.sheetnames and not annual_df.empty:
        ws = workbook["Annual Data"]
        _format_financials_sheet(ws)
    
    if "Stock Price" in workbook.sheetnames and not price_df.empty:
        ws = workbook["Stock Price"]
        ws.column_dimensions['A'].width = 20
        ws.freeze_panes = 'B2'
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15
            for cell in ws[col]:
                cell.number_format = '#,##0.00'
        ws.column_dimensions['G'].width = 18
        for cell in ws['G']:
            cell.number_format = '#,##0'

    if "Definitions" in workbook.sheetnames:
        ws = workbook["Definitions"]
        ws.column_dimensions[get_column_letter(1)].width = 25
        ws.column_dimensions[get_column_letter(2)].width = 60

    if "Annual Data" in workbook.sheetnames and not annual_df.empty:
        chart_worksheet = workbook.create_sheet("Charts")
        print("📈 Generating charts...")
        
        metrics_to_chart = [
            "Revenue (B$)", "Net Income (B$)", "EPS", "Free Cash Flow (B$)",
            "Net Profit Margin", "ROE", "Debt-to-Equity"
        ]
        
        _add_charts_to_sheet(chart_worksheet, workbook["Annual Data"], metrics_to_chart)

    workbook.save(output_filename)
    print(f"✅ Exported formatted Excel: {output_filename}")